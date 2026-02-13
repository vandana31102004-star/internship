"""
file=open("simple.txt","w")
file.write("Hello,this is a file handling exmple.")
file.close()

file=open("simple.txt","r")
content=file.read()
print(content)
file.close()

try:
    with open("missing.txt","r") as file:
     print(file.read())
except FileNotFoundError:
   print("filenot found.please check the filename.")
   


import csv
with open("C:\DS_AI internshisp\src\Day 7\Book2.csv") as file:
    reader=csv.reader(file)
    for row in reader:
        print(row)
        from openpyxl import load_workbook
        """
"""
from openpyxl import load_workbook

wb = load_workbook(
    r"C:\DS_AI internshisp\src\Day 7\Book3.xlsx"
)
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row)
    """
filename = input("Please enter the filename: ")

try:
    with open(filename, 'r') as file:
        content = file.read()
        print("File content:")
        print(content)
except FileNotFoundError:
    print("Oops! That file doesn't exist yet.")



   

